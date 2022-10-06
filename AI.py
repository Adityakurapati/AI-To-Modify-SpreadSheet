import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import pandas as pd

wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
# cell = sheet.cell(1,1)
# print(cell.value)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 4)
    # print(cell.value)
    corrected_price= cell.value * 0.9
    corrected_price_cell=sheet.cell(row,5)
    corrected_price_cell.value=corrected_price

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=5, max_col=5)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'A8')
wb.save("transactions.xlsx")
